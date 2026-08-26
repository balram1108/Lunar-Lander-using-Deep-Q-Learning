import gymnasium as gym
import torch
import torch.nn as nn
import random
import numpy as np
import csv
import pandas as pd

from collections import deque

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference


# =========================================================
# 1. DQN NEURAL NETWORK
# =========================================================

class DQN(nn.Module):

    def __init__(self):

        super().__init__()

        self.network = nn.Sequential(

            nn.Linear(8, 64),
            nn.ReLU(),

            nn.Linear(64, 64),
            nn.ReLU(),

            nn.Linear(64, 4)
        )

    def forward(self, x):

        return self.network(x)


# =========================================================
# 2. MAIN NETWORK
# =========================================================

dqn = DQN()


# =========================================================
# 3. TARGET NETWORK
# =========================================================

target_dqn = DQN()


# Start with same weights

target_dqn.load_state_dict(
    dqn.state_dict()
)


target_dqn.eval()


# =========================================================
# 4. OPTIMIZER
# =========================================================

optimizer = torch.optim.Adam(

    dqn.parameters(),

    lr=0.001
)


# =========================================================
# 5. REPLAY BUFFER
# =========================================================

replay_buffer = deque(

    maxlen=10000
)


# =========================================================
# 6. TRAINING SETTINGS
# =========================================================

gamma = 0.99

batch_size = 32


# =========================================================
# 7. EPSILON DECAY
# =========================================================

epsilon = 1.0

epsilon_min = 0.05

epsilon_decay = 0.995


# =========================================================
# 8. TARGET NETWORK SETTINGS
# =========================================================

target_update_frequency = 1000

training_step = 0


# =========================================================
# 9. LOSS FUNCTION
# =========================================================

loss_function = nn.MSELoss()


# =========================================================
# 10. ENVIRONMENT
# =========================================================

env = gym.make(

    "LunarLander-v3",

    render_mode="human"
)


# =========================================================
# 11. TRAINING REPORT FILE
# =========================================================

training_csv = "training_report_target_network.csv"


report_file = open(

    training_csv,

    "w",

    newline=""
)


csv_writer = csv.writer(
    report_file
)


csv_writer.writerow([

    "Episode",
    "Epsilon",
    "Total Reward",
    "Steps",
    "Final X",
    "Final Y",
    "Closest X To Center",
    "Left Leg",
    "Right Leg",
    "Final Reward",
    "Average Loss"

])


# =========================================================
# 12. TRAIN FOR 500 EPISODES
# =========================================================

for episode in range(500):


    state, info = env.reset()


    episode_reward = 0

    episode_losses = []

    closest_x_to_center = abs(
        state[0]
    )

    final_reward = 0

    episode_epsilon = epsilon


    # =====================================================
    # RUN MAXIMUM 500 STEPS
    # =====================================================

    for step_number in range(500):


        # -------------------------------------------------
        # STATE -> TENSOR
        # -------------------------------------------------

        state_tensor = torch.tensor(

            state,

            dtype=torch.float32
        )


        # -------------------------------------------------
        # MAIN DQN Q-VALUES
        # -------------------------------------------------

        q_values = dqn(
            state_tensor
        )


        # =================================================
        # EPSILON-GREEDY ACTION
        # =================================================

        if random.random() < epsilon:


            # Explore

            action = env.action_space.sample()


        else:


            # Exploit

            action = torch.argmax(
                q_values
            ).item()


        # =================================================
        # ENVIRONMENT STEP
        # =================================================

        next_state, reward, terminated, truncated, info = env.step(
            action
        )


        done = terminated or truncated


        # =================================================
        # REPORT INFORMATION
        # =================================================

        episode_reward += reward

        final_reward = reward


        current_x_distance = abs(
            next_state[0]
        )


        if current_x_distance < closest_x_to_center:

            closest_x_to_center = current_x_distance


        # =================================================
        # STORE EXPERIENCE
        # =================================================

        experience = (

            state,
            action,
            reward,
            next_state,
            done

        )


        replay_buffer.append(
            experience
        )


        # =================================================
        # START TRAINING AFTER 32 EXPERIENCES
        # =================================================

        if len(replay_buffer) >= batch_size:


            # ---------------------------------------------
            # RANDOMLY SAMPLE 32 EXPERIENCES
            # ---------------------------------------------

            batch = random.sample(

                replay_buffer,

                batch_size
            )


            states, actions, rewards, next_states, dones = zip(
                *batch
            )


            # ---------------------------------------------
            # CONVERT TO TENSORS
            # ---------------------------------------------

            states_tensor = torch.tensor(

                np.array(states),

                dtype=torch.float32
            )


            actions_tensor = torch.tensor(

                actions,

                dtype=torch.long
            )


            rewards_tensor = torch.tensor(

                rewards,

                dtype=torch.float32
            )


            next_states_tensor = torch.tensor(

                np.array(next_states),

                dtype=torch.float32
            )


            dones_tensor = torch.tensor(

                dones,

                dtype=torch.float32
            )


            # =================================================
            # MAIN NETWORK PREDICTION
            # =================================================

            all_q_values = dqn(
                states_tensor
            )


            predicted_q_values = all_q_values.gather(

                1,

                actions_tensor.unsqueeze(1)

            ).squeeze(1)


            # =================================================
            # TARGET NETWORK
            # =================================================

            with torch.no_grad():


                next_q_values = target_dqn(
                    next_states_tensor
                )


                best_next_q_values = next_q_values.max(
                    dim=1
                ).values


                # ---------------------------------------------
                # TD TARGET
                # ---------------------------------------------

                targets = (

                    rewards_tensor

                    +

                    gamma
                    * best_next_q_values
                    * (1 - dones_tensor)

                )


            # =================================================
            # LOSS
            # =================================================

            loss = loss_function(

                predicted_q_values,

                targets
            )


            episode_losses.append(
                loss.item()
            )


            # =================================================
            # UPDATE MAIN NETWORK WEIGHTS
            # =================================================

            optimizer.zero_grad()

            loss.backward()

            optimizer.step()


            training_step += 1


            # =================================================
            # UPDATE TARGET NETWORK
            # =================================================

            if training_step % target_update_frequency == 0:


                target_dqn.load_state_dict(
                    dqn.state_dict()
                )


                print(
                    "TARGET NETWORK UPDATED"
                )


        # =================================================
        # EPISODE FINISHED
        # =================================================

        if done:

            break


        # State 2 becomes current state

        state = next_state


    # =====================================================
    # END OF TRAINING EPISODE
    # =====================================================

    final_x = next_state[0]

    final_y = next_state[1]

    left_leg = next_state[6]

    right_leg = next_state[7]


    if len(episode_losses) > 0:

        average_loss = (

            sum(episode_losses)

            /

            len(episode_losses)

        )

    else:

        average_loss = 0


    # =====================================================
    # PRINT TRAINING REPORT
    # =====================================================

    print()

    print("==============================")

    print("TRAINING EPISODE:", episode)

    print("==============================")


    print(
        "Epsilon:",
        round(episode_epsilon, 4)
    )


    print(
        "Total Reward:",
        round(episode_reward, 2)
    )


    print(
        "Steps:",
        step_number + 1
    )


    print(
        "Final Reward:",
        round(final_reward, 2)
    )


    print(
        "Average Loss:",
        round(average_loss, 4)
    )


    print("==============================")


    # =====================================================
    # SAVE TRAINING RESULT
    # =====================================================

    csv_writer.writerow([

        episode,

        episode_epsilon,

        episode_reward,

        step_number + 1,

        final_x,

        final_y,

        closest_x_to_center,

        left_leg,

        right_leg,

        final_reward,

        average_loss

    ])


    report_file.flush()


    # =====================================================
    # DECAY EPSILON
    # =====================================================

    epsilon = max(

        epsilon_min,

        epsilon * epsilon_decay

    )


# =========================================================
# TRAINING FINISHED
# =========================================================

report_file.close()


print()

print("====================================")

print("500 TRAINING EPISODES COMPLETE")

print("====================================")

print()


# =========================================================
# 13. EVALUATION
#
# EPSILON = 0
# NO TRAINING
# =========================================================

dqn.eval()


test_episodes = 100


successful_landings = 0


evaluation_results = []


# =========================================================
# RUN 100 TEST EPISODES
# =========================================================

for test_episode in range(test_episodes):


    state, info = env.reset()


    total_test_reward = 0


    final_test_reward = 0


    # =====================================================
    # TEST EPISODE
    # =====================================================

    for step_number in range(500):


        state_tensor = torch.tensor(

            state,

            dtype=torch.float32
        )


        # =================================================
        # NO TRAINING
        # =================================================

        with torch.no_grad():


            q_values = dqn(
                state_tensor
            )


        # =================================================
        # EPSILON = 0
        #
        # ALWAYS PICK HIGHEST Q-VALUE
        # =================================================

        action = torch.argmax(
            q_values
        ).item()


        # =================================================
        # TAKE ACTION
        # =================================================

        next_state, reward, terminated, truncated, info = env.step(
            action
        )


        total_test_reward += reward


        final_test_reward = reward


        done = terminated or truncated


        # =================================================
        # CHECK SUCCESS
        # =================================================

        if done:


            if reward == 100:

                successful_landings += 1


            break


        state = next_state


    # =====================================================
    # SAVE TEST RESULT
    # =====================================================

    evaluation_results.append([

        test_episode,

        total_test_reward,

        step_number + 1,

        next_state[0],

        next_state[1],

        next_state[6],

        next_state[7],

        final_test_reward

    ])


    print(

        "TEST EPISODE:",

        test_episode,

        "| Reward:",

        round(total_test_reward, 2)

    )


# =========================================================
# 14. CREATE EVALUATION CSV
# =========================================================

evaluation_csv = "evaluation_report.csv"


with open(

    evaluation_csv,

    "w",

    newline=""

) as evaluation_file:


    evaluation_writer = csv.writer(
        evaluation_file
    )


    evaluation_writer.writerow([

        "Test Episode",

        "Total Reward",

        "Steps",

        "Final X",

        "Final Y",

        "Left Leg",

        "Right Leg",

        "Final Reward"

    ])


    evaluation_writer.writerows(
        evaluation_results
    )


# =========================================================
# 15. FINAL EVALUATION RESULTS
# =========================================================

test_rewards = [

    row[1]

    for row in evaluation_results

]


average_test_reward = (

    sum(test_rewards)

    /

    len(test_rewards)

)


success_rate = (

    successful_landings

    /

    test_episodes

) * 100


print()

print("====================================")

print("FINAL EVALUATION")

print("====================================")


print(
    "Test Episodes:",
    test_episodes
)


print(
    "Successful Landings:",
    successful_landings
)


print(
    "Success Rate:",
    round(success_rate, 2),
    "%"
)


print(
    "Average Test Reward:",
    round(average_test_reward, 2)
)


print("====================================")


# =========================================================
# 16. CREATE EXCEL REPORT
# =========================================================

training_df = pd.read_csv(
    training_csv
)


evaluation_df = pd.read_csv(
    evaluation_csv
)


# =========================================================
# ADD SMOOTH TRAINING REWARD
# =========================================================

training_df["25 Episode Average Reward"] = (

    training_df["Total Reward"]

    .rolling(

        window=25,

        min_periods=1

    )

    .mean()

)


# =========================================================
# ADD SMOOTH TRAINING LOSS
# =========================================================

training_df["25 Episode Average Loss"] = (

    training_df["Average Loss"]

    .rolling(

        window=25,

        min_periods=1

    )

    .mean()

)


# =========================================================
# ADD SMOOTH TEST REWARD
# =========================================================

evaluation_df["10 Episode Average Reward"] = (

    evaluation_df["Total Reward"]

    .rolling(

        window=10,

        min_periods=1

    )

    .mean()

)


# =========================================================
# SAVE BOTH SHEETS TO EXCEL
# =========================================================

excel_filename = "lunar_lander_results.xlsx"


with pd.ExcelWriter(

    excel_filename,

    engine="openpyxl"

) as writer:


    training_df.to_excel(

        writer,

        sheet_name="Training",

        index=False

    )


    evaluation_df.to_excel(

        writer,

        sheet_name="Evaluation",

        index=False

    )


# =========================================================
# 17. ADD GRAPHS TO EXCEL
# =========================================================

workbook = load_workbook(
    excel_filename
)


# =========================================================
# TRAINING GRAPH
# =========================================================

training_sheet = workbook["Training"]


training_chart = LineChart()


training_chart.title = (
    "Training - 25 Episode Average Reward"
)


training_chart.y_axis.title = (
    "Average Reward"
)


training_chart.x_axis.title = (
    "Episode"
)


training_reward_column = (

    training_df.columns.get_loc(
        "25 Episode Average Reward"
    )

    + 1

)


training_data = Reference(

    training_sheet,

    min_col=training_reward_column,

    min_row=1,

    max_row=training_sheet.max_row

)


training_categories = Reference(

    training_sheet,

    min_col=1,

    min_row=2,

    max_row=training_sheet.max_row

)


training_chart.add_data(

    training_data,

    titles_from_data=True

)


training_chart.set_categories(
    training_categories
)


training_chart.height = 10

training_chart.width = 20


training_sheet.add_chart(

    training_chart,

    "N2"

)


# =========================================================
# EVALUATION GRAPH
# =========================================================

evaluation_sheet = workbook["Evaluation"]


evaluation_chart = LineChart()


evaluation_chart.title = (
    "Evaluation - Reward"
)


evaluation_chart.y_axis.title = (
    "Reward"
)


evaluation_chart.x_axis.title = (
    "Test Episode"
)


evaluation_data = Reference(

    evaluation_sheet,

    min_col=2,

    min_row=1,

    max_row=evaluation_sheet.max_row

)


evaluation_categories = Reference(

    evaluation_sheet,

    min_col=1,

    min_row=2,

    max_row=evaluation_sheet.max_row

)


evaluation_chart.add_data(

    evaluation_data,

    titles_from_data=True

)


evaluation_chart.set_categories(
    evaluation_categories
)


evaluation_chart.height = 10

evaluation_chart.width = 20


evaluation_sheet.add_chart(

    evaluation_chart,

    "K2"

)


# =========================================================
# SAVE EXCEL
# =========================================================

workbook.save(
    excel_filename
)


# =========================================================
# CLOSE ENVIRONMENT
# =========================================================

env.close()


# =========================================================
# FINISHED
# =========================================================

print()

print("====================================")

print("ALL DONE")

print("====================================")


print(
    "Training CSV:",
    training_csv
)


print(
    "Evaluation CSV:",
    evaluation_csv
)


print(
    "Excel report:",
    excel_filename
)


print("====================================")